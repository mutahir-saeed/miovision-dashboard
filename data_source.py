"""
Layer 1 — INGESTION.

This is the ONLY module that knows *where* the data lives or *how* it is shaped
on disk. Everything downstream (aggregations, dashboard) consumes the single tidy
"long" DataFrame returned by `load_long_table()` and never touches Excel again.

To switch to Google Sheets later: write a `load_long_table_from_gsheet()` that
returns a DataFrame with the SAME columns, and point the app at it. Nothing else
in the project changes.
"""

from __future__ import annotations
import io
import re
import pandas as pd
import openpyxl

# Canonical column names used everywhere downstream.
COL_STUDENT = "student"
COL_PROJECT = "project"
COL_TASK = "task"
COL_HOURS = "hours"
COL_SIGNALS = "signals"
COL_WEEK = "week"
COL_DATE = "date"
COL_NOTES = "notes"

UNSPECIFIED_TASK = "(Unspecified)"
REFERENCE_SHEET = "Reference"


def _open_workbook(source):
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    return openpyxl.load_workbook(source, data_only=True)


def _match_header(headers: dict[str, int], *needles: str, exclude: str | None = None) -> int | None:
    """Find the 1-based column index whose header contains all needles (case-insensitive)."""
    for name, idx in headers.items():
        low = name.lower()
        if exclude and exclude in low:
            continue
        if all(n in low for n in needles):
            return idx
    return None


def _read_headers(ws) -> dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c
    return headers


def _clean_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_reference(path: str) -> dict:
    """Return the project -> [tasks] structure and holiday list from the Reference tab."""
    wb = _open_workbook(path)
    ws = wb[REFERENCE_SHEET]
    projects: dict[str, list[str]] = {}
    # Row 1 = project headers; rows below = tasks, until a blank or the 'Public Holidays' marker.
    for c in range(1, ws.max_column + 1):
        proj = _clean_str(ws.cell(row=1, column=c).value)
        if not proj:
            continue
        tasks = []
        for r in range(2, 11):  # tasks sit just under the header
            t = _clean_str(ws.cell(row=r, column=c).value)
            if t:
                tasks.append(t)
        projects[proj] = tasks
    holidays = []
    for r in range(13, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if isinstance(d, (pd.Timestamp,)) or hasattr(d, "year"):
            holidays.append(pd.to_datetime(d))
    return {"projects": projects, "holidays": holidays}


def load_long_table(path: str) -> pd.DataFrame:
    """
    Collapse every student tab into one tidy long table.

    One row = one logged day for one student. Returns columns:
    student, project, task, hours, signals, week, date, notes
    Only rows with real work (hours or signals present) are kept.
    """
    wb = _open_workbook(path)
    records = []

    for name in wb.sheetnames:
        if name == REFERENCE_SHEET:
            continue
        ws = wb[name]
        h = _read_headers(ws)

        c_proj = _match_header(h, "project")
        c_task = _match_header(h, "task")
        c_hours = _match_header(h, "duration")
        c_signals = _match_header(h, "signal", exclude="time")  # "# of Signals" / "No. of Signals"
        c_week = _match_header(h, "kw") or _match_header(h, "week")
        c_date = _match_header(h, "date")
        c_notes = _match_header(h, "notes")

        for r in range(2, ws.max_row + 1):
            hours = _to_float(ws.cell(row=r, column=c_hours).value) if c_hours else None
            signals = _to_float(ws.cell(row=r, column=c_signals).value) if c_signals else None
            proj = _clean_str(ws.cell(row=r, column=c_proj).value) if c_proj else None

            # Keep a row only if it represents real logged work.
            if hours is None and signals is None and proj is None:
                continue
            if hours is None and signals is None:
                continue

            task = _clean_str(ws.cell(row=r, column=c_task).value) if c_task else None
            records.append({
                COL_STUDENT: name,
                COL_PROJECT: proj or UNSPECIFIED_TASK,
                COL_TASK: task or UNSPECIFIED_TASK,
                COL_HOURS: hours or 0.0,
                COL_SIGNALS: signals,  # may be None for non-signal tasks
                COL_WEEK: _to_float(ws.cell(row=r, column=c_week).value) if c_week else None,
                COL_DATE: pd.to_datetime(ws.cell(row=r, column=c_date).value, errors="coerce") if c_date else pd.NaT,
                COL_NOTES: _clean_str(ws.cell(row=r, column=c_notes).value) if c_notes else None,
            })

    df = pd.DataFrame.from_records(records)
    if not df.empty:
        df[COL_WEEK] = df[COL_WEEK].astype("Int64")
    return df


if __name__ == "__main__":
    df = load_long_table("data.xlsx")
    print(f"Loaded {len(df)} logged rows from {df[COL_STUDENT].nunique()} students")
    print(df.head(10).to_string())
    print("\nProjects:", sorted(df[COL_PROJECT].unique()))
